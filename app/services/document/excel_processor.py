from app.processors.base_processor import BaseDocumentProcessor
from typing import List, Dict, Any, Optional
import pandas as pd
import logging
from openpyxl import load_workbook
import re
from datetime import datetime

class ExcelProcessor(BaseDocumentProcessor):
    """Excel 파일 처리기"""
    
    def process(self, file_path: str) -> List[str]:
        """
        Excel 파일에서 향상된 텍스트 추출
        
        Args:
            file_path (str): Excel 파일 경로
            
        Returns:
            List[str]: 추출된 텍스트 청크 목록
        """
        try:
            chunks = []
            
            # 워크북 메타데이터 추출
            metadata_info = self._extract_workbook_metadata(file_path)
            if metadata_info:
                chunks.append(f"[문서 정보]\n{metadata_info}")
            
            # openpyxl로 워크북 로드 (더 많은 정보 접근을 위해)
            try:
                wb = load_workbook(file_path, data_only=False)
                wb_data_only = load_workbook(file_path, data_only=True)
                
                # 워크북 구조 정보
                structure_info = self._extract_workbook_structure(wb)
                if structure_info:
                    chunks.append(f"[워크북 구조]\n{structure_info}")
                
                # 각 시트 처리
                for sheet_name in wb.sheetnames:
                    try:
                        # 시트 정보 청크
                        sheet_chunks = self._process_enhanced_sheet(
                            wb[sheet_name], 
                            wb_data_only[sheet_name] if sheet_name in wb_data_only.sheetnames else None,
                            sheet_name
                        )
                        chunks.extend(sheet_chunks)
                    except Exception as sheet_error:
                        logging.error(f"시트 '{sheet_name}' 처리 중 오류: {sheet_error}")
                        chunks.append(f"시트 '{sheet_name}' 처리 중 오류가 발생했습니다.")
                
            except Exception as openpyxl_error:
                logging.warning(f"openpyxl로 처리 실패, pandas로 대체: {openpyxl_error}")
                # 기본 pandas 처리로 대체
                chunks.extend(self._fallback_pandas_processing(file_path))
            
            # 결과 정리
            if not chunks:
                return ["Excel 파일에서 내용을 추출할 수 없습니다."]
                
            logging.info(f"Excel 처리 완료: {file_path}, {len(chunks)}개 청크 생성")
            return chunks
            
        except Exception as e:
            logging.error(f"Excel 처리 중 오류 발생: {e}")
            return [f"Excel 파일 처리 중 오류가 발생했습니다: {str(e)}"]
    
    def _extract_workbook_metadata(self, file_path: str) -> str:
        """워크북 메타데이터 추출"""
        metadata_parts = []
        
        try:
            wb = load_workbook(file_path)
            props = wb.properties
            
            if props.title:
                metadata_parts.append(f"제목: {props.title}")
            if props.creator:
                metadata_parts.append(f"작성자: {props.creator}")
            if props.subject:
                metadata_parts.append(f"주제: {props.subject}")
            if props.description:
                metadata_parts.append(f"설명: {props.description}")
            if props.keywords:
                metadata_parts.append(f"키워드: {props.keywords}")
            if props.created:
                metadata_parts.append(f"생성일: {props.created.strftime('%Y-%m-%d %H:%M:%S')}")
            if props.modified:
                metadata_parts.append(f"수정일: {props.modified.strftime('%Y-%m-%d %H:%M:%S')}")
            if props.lastModifiedBy:
                metadata_parts.append(f"최종 수정자: {props.lastModifiedBy}")
                
        except Exception as e:
            logging.warning(f"워크북 메타데이터 추출 중 오류: {e}")
        
        return "\n".join(metadata_parts) if metadata_parts else ""
    
    def _extract_workbook_structure(self, wb) -> str:
        """워크북 구조 정보 추출"""
        structure_parts = []
        
        try:
            # 시트 목록
            structure_parts.append(f"시트 수: {len(wb.sheetnames)}")
            structure_parts.append(f"시트 목록: {', '.join(wb.sheetnames)}")
            
            # 정의된 이름(Named Ranges) 
            if wb.defined_names:
                defined_names = []
                for name in wb.defined_names:
                    try:
                        defined_names.append(f"{name.name}: {name.value}")
                    except:
                        defined_names.append(name.name)
                
                if defined_names:
                    structure_parts.append(f"정의된 이름: {', '.join(defined_names[:10])}")  # 처음 10개만
            
        except Exception as e:
            logging.warning(f"워크북 구조 추출 중 오류: {e}")
        
        return "\n".join(structure_parts) if structure_parts else ""
    
    def _process_enhanced_sheet(self, ws, ws_data_only, sheet_name: str) -> List[str]:
        """향상된 시트 처리"""
        chunks = []
        
        try:
            # 시트 기본 정보
            sheet_info_parts = [f"시트명: {sheet_name}"]
            
            # 시트 크기 정보
            if ws.max_row > 0 and ws.max_column > 0:
                sheet_info_parts.append(f"크기: {ws.max_row}행 x {ws.max_column}열")
                sheet_info_parts.append(f"사용 범위: A1:{ws.max_column}열{ws.max_row}행")
            
            chunks.append(f"[시트: {sheet_name}]\n" + "\n".join(sheet_info_parts))
            
            # 차트 정보 추출
            chart_info = self._extract_chart_info(ws, sheet_name)
            if chart_info:
                chunks.append(chart_info)
            
            # 데이터를 DataFrame으로 변환하여 처리
            data_chunks = self._extract_sheet_data(ws, ws_data_only, sheet_name)
            chunks.extend(data_chunks)
            
            # 수식 정보 추출 (중요한 수식만)
            formula_info = self._extract_formula_info(ws, sheet_name)
            if formula_info:
                chunks.append(formula_info)
                
        except Exception as e:
            logging.error(f"시트 '{sheet_name}' 향상 처리 중 오류: {e}")
            chunks.append(f"시트 '{sheet_name}' 처리 중 오류 발생")
        
        return chunks
    
    def _extract_chart_info(self, ws, sheet_name: str) -> Optional[str]:
        """차트 정보 추출"""
        if not ws._charts:
            return None
        
        chart_info_parts = [f"차트 정보 (시트: {sheet_name}):"]
        
        for i, chart in enumerate(ws._charts):
            try:
                chart_type = type(chart).__name__
                chart_info_parts.append(f"차트 {i+1}: {chart_type}")
                
                # 차트 제목
                if hasattr(chart, 'title') and chart.title:
                    chart_info_parts.append(f"  제목: {chart.title}")
                
                # 데이터 범위 (가능한 경우)
                if hasattr(chart, 'series') and chart.series:
                    for j, series in enumerate(chart.series[:3]):  # 처음 3개 시리즈만
                        try:
                            if hasattr(series, 'title') and series.title:
                                chart_info_parts.append(f"  시리즈 {j+1}: {series.title}")
                        except:
                            pass
                            
            except Exception as e:
                logging.warning(f"차트 {i+1} 정보 추출 중 오류: {e}")
        
        return "\n".join(chart_info_parts) if len(chart_info_parts) > 1 else None
    
    def _extract_sheet_data(self, ws, ws_data_only, sheet_name: str) -> List[str]:
        """시트 데이터 추출 및 청킹"""
        chunks = []
        
        try:
            # 데이터가 있는 실제 범위 찾기
            data_range = self._find_data_range(ws)
            if not data_range:
                return [f"시트 '{sheet_name}'에 데이터가 없습니다."]
            
            min_row, max_row, min_col, max_col = data_range
            
            # 헤더 추출 (첫 번째 행이 헤더로 가정)
            headers = []
            for col in range(min_col, max_col + 1):
                cell_value = ws.cell(row=min_row, column=col).value
                if cell_value is not None:
                    headers.append(str(cell_value))
                else:
                    headers.append(f"Column_{col}")
            
            # 헤더 정보 청크
            if headers:
                chunks.append(f"[{sheet_name} - 헤더]\n" + " | ".join(headers))
            
            # 데이터 청킹 (적절한 크기로)
            chunk_size = 25  # 행 단위
            current_chunk_data = []
            
            for row_idx in range(min_row + 1, max_row + 1):
                row_data = []
                has_data = False
                
                for col_idx in range(min_col, max_col + 1):
                    # 값 우선, 수식이 있다면 계산된 값 사용
                    cell_value = None
                    if ws_data_only:
                        cell_value = ws_data_only.cell(row=row_idx, column=col_idx).value
                    
                    if cell_value is None:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                    
                    if cell_value is not None:
                        row_data.append(str(cell_value))
                        has_data = True
                    else:
                        row_data.append("")
                
                if has_data:
                    current_chunk_data.append(f"행{row_idx}: " + " | ".join(row_data))
                
                # 청크 크기에 도달하면 저장
                if len(current_chunk_data) >= chunk_size:
                    chunk_text = f"[{sheet_name} - 데이터 (행 {row_idx-len(current_chunk_data)+1}-{row_idx})]\n"
                    chunk_text += "\n".join(current_chunk_data)
                    chunks.append(chunk_text)
                    current_chunk_data = []
            
            # 남은 데이터 처리
            if current_chunk_data:
                chunk_text = f"[{sheet_name} - 데이터 (마지막)]\n"
                chunk_text += "\n".join(current_chunk_data)
                chunks.append(chunk_text)
        
        except Exception as e:
            logging.error(f"시트 '{sheet_name}' 데이터 추출 중 오류: {e}")
            chunks.append(f"시트 '{sheet_name}' 데이터 추출 실패")
        
        return chunks
    
    def _find_data_range(self, ws) -> Optional[tuple]:
        """실제 데이터가 있는 범위 찾기"""
        try:
            min_row = min_col = float('inf')
            max_row = max_col = 0
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        min_row = min(min_row, cell.row)
                        max_row = max(max_row, cell.row)
                        min_col = min(min_col, cell.column)
                        max_col = max(max_col, cell.column)
            
            if min_row == float('inf'):
                return None
                
            return (min_row, max_row, min_col, max_col)
            
        except Exception as e:
            logging.error(f"데이터 범위 찾기 중 오류: {e}")
            return None
    
    def _extract_formula_info(self, ws, sheet_name: str) -> Optional[str]:
        """중요한 수식 정보 추출"""
        formula_info_parts = []
        formula_count = 0
        
        try:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:  # 수식 셀
                        formula_count += 1
                        if formula_count <= 10:  # 처음 10개 수식만 기록
                            cell_ref = f"{cell.coordinate}"
                            formula_info_parts.append(f"{cell_ref}: {cell.value}")
            
            if formula_info_parts:
                result_parts = [f"수식 정보 (시트: {sheet_name}, 총 {formula_count}개):"]
                result_parts.extend(formula_info_parts)
                if formula_count > 10:
                    result_parts.append(f"... 외 {formula_count - 10}개 수식")
                return "\n".join(result_parts)
                
        except Exception as e:
            logging.warning(f"수식 정보 추출 중 오류: {e}")
        
        return None
    
    def _fallback_pandas_processing(self, file_path: str) -> List[str]:
        """pandas를 사용한 대체 처리"""
        chunks = []
        
        try:
            df_dict = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            
            for sheet_name, df in df_dict.items():
                if df.empty:
                    chunks.append(f"시트 '{sheet_name}'는 비어있습니다.")
                    continue
                
                # 시트 정보
                sheet_info = f"시트: {sheet_name}\n크기: {len(df)}행 x {len(df.columns)}열"
                chunks.append(sheet_info)
                
                # 컬럼 정보
                columns_info = f"컬럼: {', '.join(df.columns.astype(str))}"
                chunks.append(columns_info)
                
                # 데이터 처리
                if len(df) > 30:
                    chunk_size = 30
                    for i in range(0, len(df), chunk_size):
                        chunk_df = df.iloc[i:i+chunk_size]
                        chunk_text = self._process_dataframe(chunk_df, sheet_name, i)
                        chunks.append(chunk_text)
                else:
                    chunk_text = self._process_dataframe(df, sheet_name, 0)
                    chunks.append(chunk_text)
            
        except Exception as e:
            logging.error(f"pandas 대체 처리 중 오류: {e}")
            chunks.append("Excel 파일 처리 중 오류가 발생했습니다.")
        
        return chunks
    
    def _process_dataframe(self, df: pd.DataFrame, sheet_name: str, start_index: int) -> str:
        """
        DataFrame을 텍스트로 변환
        
        Args:
            df (pd.DataFrame): 변환할 DataFrame
            sheet_name (str): 시트 이름
            start_index (int): 시작 인덱스
            
        Returns:
            str: 변환된 텍스트
        """
        rows_text = []
        
        # 헤더 추가
        header = f"시트 '{sheet_name}' 데이터 (행 {start_index+1}-{start_index+len(df)}):"
        rows_text.append(header)
        
        # 각 행을 처리
        for idx, row in df.iterrows():
            # NaN 값 제거
            row_clean = {col: val for col, val in row.items() if pd.notna(val)}
            
            # 행 번호와 함께 텍스트 형식으로 변환
            row_text = f"행 {idx+1}: " + " | ".join([f"{col}: {val}" for col, val in row_clean.items()])
            rows_text.append(row_text)
        
        return "\n".join(rows_text)